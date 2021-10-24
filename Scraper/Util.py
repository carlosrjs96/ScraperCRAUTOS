# Import Libraries
from selenium import webdriver
import random
import time
import json
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os
from datetime import datetime
#-----------------------------------------------------------------------
def selenium_WebDriver_config():
    # Navigation options
    options =  webdriver.ChromeOptions()
    options.add_argument('--incognito')# direct access to browser incognito mode
    options.add_argument('--start-maximized')
    options.add_argument('--disable-extensions')
    #https://stackoverflow.com/questions/53039551/selenium-webdriver-modifying-navigator-webdriver-flag-to-prevent-selenium-detec/53040904#53040904
    #https://stackoverflow.com/questions/62490495/how-to-change-the-user-agent-using-selenium-and-python/62491863#62491863
    options.add_argument("--disable-blink-features")
    
    # For ChromeDriver version 79.0.3945.16 or over
    options.add_argument("--disable-blink-features=AutomationControlled") 
    
    # For older ChromeDriver under version 79.0.3945.16
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)

    driver_path = './chromedriver.exe'
    driver = webdriver.Chrome(driver_path, options=options)
    #driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)

    #Setting up Chrome/89.0.4389.114 as useragent
    driver.execute_cdp_cmd('Network.setUserAgentOverride', {"userAgent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.72 Safari/537.36'})
    #                                                                     Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36
    
    #Remove navigator.webdriver Flag using JavaScript
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    #print(driver.execute_script("return navigator.userAgent;"))
    
    # Start screen
    driver.maximize_window() 

    #driver.get('https://www.httpbin.org/headers')
    #staticDelay(0,10,11)
    
    return driver
#-----------------------------------------------------------------------
def staticDelay(num1,num2):
    delay = random.uniform(num1,num2)
    #print("\nSleep : "+str(delay)+" seconds")
    time.sleep(delay)# DELAY
#-----------------------------------------------------------------------
def getContentBodyTable(driver, before_XPath, row_Gap, column_Gap):
        
    # Get Rows Size
    num_rows = len (driver.find_elements_by_xpath(before_XPath)) + row_Gap
    print("Rows in table are " + repr(num_rows))
    
    # Get Columns Size
    num_cols = len (driver.find_elements_by_xpath(before_XPath +'[1]/td')) + column_Gap
    print("Columns in table are " + repr(num_cols))
    # XPath Structure
    before_XPath = before_XPath + '['
    aftertd_XPath = ']/td['
    aftertr_XPath = ']'
    # Load Data Table
    table = []
    for t_row in range(1,(num_rows)): #tr[7]/td[1]/strong --> tr[7]/td[1]/div/span
            data = []
            for t_column in range(1, (num_cols+1)):
                FinalXPath = before_XPath + str(t_row) + aftertd_XPath + str(t_column) + aftertr_XPath
                cell_text = driver.find_element_by_xpath(FinalXPath) 
                data.append(cell_text)
                #print("\nDato > "+str(cell_text))
            table.append(data)
    return table
#-----------------------------------------------------------------------
def getContentHeadTable(driver, before_XPath):
    # Get Columns Size
    num_cols = len (driver.find_elements_by_xpath(before_XPath))
    #print("Columns in table are " + repr(num_cols))

    # XPath Structure
    before_XPath = before_XPath + '['
    afterth_XPath = ']'

    # Load Data Table
    table = []
    for t_column in range(1, (num_cols+1)):
        FinalXPath = before_XPath + str(t_column) + afterth_XPath
        cell_text = driver.find_element_by_xpath(FinalXPath) 
        table.append(cell_text)
        #print("\nDato > "+str(cell_text))

    return table
#-----------------------------------------------------------------------
def find_element_in_list(element, list_element):
    try:
        index_element = list_element.index(element)
        return index_element
    except ValueError:
        return None
#-----------------------------------------------------------------------
def getDymanicContentBodyTable(driver, before_XPath, row_Gap, column_Gap):
        
    # Get Rows Size
    num_rows = len (driver.find_elements_by_xpath(before_XPath)) + row_Gap
    #print("Rows in table are " + repr(num_rows))
    
    # XPath Structure
    before_XPath = before_XPath + '['
    aftertd_XPath = ']/td['
    aftertr_XPath = ']'
    #/html/body/section[1]/div/div/div[1]/div[1]/table/tbody/tr[5]/td/div[1]/table/tbody/tr[1]/td
    # Load Data Table
    table = []
    for t_row in range(1,(num_rows+1)): #tr[7]/td[1]/strong --> tr[7]/td[1]/div/span
        data = []
        # Get Columns Size
        num_cols = len(driver.find_elements_by_xpath(before_XPath +str(t_row)+']/td')) + column_Gap
        #print("Columns in table are " + repr(num_cols))
        for t_column in range(num_cols):
            FinalXPath = before_XPath + str(t_row) + aftertd_XPath + str(t_column+1) + aftertr_XPath
            cell_text = driver.find_element_by_xpath(FinalXPath) 
            data.append(cell_text)
            #print("\nDato > "+str(cell_text))
        table.append(data)
    
    return table
#-----------------------------------------------------------------------
def load_Json(path:str) -> json:
    # Load Json file
    with open(path, 'r') as j:
        _json = json.load(j)
    print(f'Json File: {path} (loaded)') 
    return _json
#-----------------------------------------------------------------------
def createExcel(workbook_name  = 'Datos.xlsx', title = 'Autos', headers = []):
    #headers       = ['Company','Address','Tel','Web']
    wb = Workbook()
    page = wb.active
    page.title = title
    page.append(headers) # write the headers to the first line
    wb.save(filename = workbook_name)
#-----------------------------------------------------------------------
def appendRowToExcel( workbook_name  = 'Datos',row = []):
    # New data to write:
    #info = [
    #    empresa.Nombre    ,
    #    empresa.Direccion ,
    #    empresa.Localidad ,
    #    empresa.Provincia ,
    #    empresa.Telefono  ,
    #    empresa.Correo    ,
    #    empresa.Actividad ,
    #    empresa.SitioWeb  ,
    #    empresa.Url       
    #]
    # Load workbook
    wb = load_workbook(workbook_name)
    page = wb.active
    page.append(row)
    wb.save(filename=workbook_name)
#-----------------------------------------------------------------------
def create_Directory(path):
    # Create target directory & all intermediate directories if don't exists
    if not os.path.exists(path):
        os.makedirs(path)
        print(f'Directory: {path} (created)')
    else:       
        print(f'Directory: {path} (already exists)') 
#-----------------------------------------------------------------------
def get_Timestamp() -> str:
    timestamp = datetime.now()  # current date and time
    timestamp = datetime.fromtimestamp(timestamp.timestamp()) # convert to datetime
    timestamp = timestamp.strftime("%m/%d/%Y %H:%M") # convert timestamp to string in dd-mm-yyyy HH:MM:SS
    return timestamp
#-----------------------------------------------------------------------